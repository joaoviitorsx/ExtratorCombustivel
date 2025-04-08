import pandas as pd
import os
from PySide6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def formatar_cnpj(cnpj):
    """Aplica a máscara de CNPJ"""
    cnpj = ''.join(filter(str.isdigit, str(cnpj)))
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj


def converter_planilha(caminho_entrada, caminho_saida):
    try:
        df = pd.read_excel(caminho_entrada, dtype=str)

        df_formatado = pd.DataFrame({
            "Tipo Doc.": ["NFSe"] * len(df),
            "Número": df["Numero"],
            "Código de Verificação": "",
            "Competência": df["Mes competencia"].astype(str).str.zfill(2) + "/" + df["Ano competencia"].astype(str),
            "Data": pd.to_datetime(df["Data"], dayfirst=True, errors='coerce').dt.strftime("%d/%m/%Y"),
            "Vencimento": "",
            "Número RPS": "",
            "Série RPS": "",
            "Tipo RPS": "",
            "Natureza da Operação": "",
            "Regime Especial Tributação": "",
            "Operação Simples Nacional": "",
            "Incentivador Cultural": "",
            "Item da Lista": df["Item servico"],
            "CNAE": "",
            "ART": "",
            "Código Obra": "",
            "Número Empenho": "",
            "Discriminação dos Serviços": "",
            "Valor dos Serviços": df["Valor faturado"],
            "Deduções Permitidas em Lei": "",
            "Desconto Condicionado": "",
            "Desconto Incondicionado": "",
            "Retenções Federais": "",
            "Outras Retenções": "",
            "PIS": "",
            "COFINS": "",
            "IRRF": "",
            "CSLL": "",
            "INSS": "",
            "Base de Cálculo": df["Base de calculo"],
            "Alíquota": df["Aliquota"],
            "Local da Prestação": "",
            "ISS Retido": df["Iss retido"],
            "Valor do ISS": df["Valor iss"],
            "Valor Líquido": "",
            "Status Doc.": "Não Incidência",
            "Inscrição Prestador": "",
            "CPF/CNPJ Prestador": df["Doc prestador"].apply(formatar_cnpj),
            "Razão Social/Nome do Prestador": df["Nome prestador"],
            "Escrituração": "Não Incidência",
            "Origem": "Prestador",
            "Status Aceite": "Aceita",
        })

        with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
            df_formatado.to_excel(writer, index=False, sheet_name="Planilha Convertida")

        aplicar_formatacao_excel(caminho_saida)

        resposta = QMessageBox.question(
            None, "Conversão concluída",
            "\nDeseja abrir a planilha agora?",
            QMessageBox.Yes | QMessageBox.No
        )
        if resposta == QMessageBox.Yes:
            os.startfile(caminho_saida)

    except Exception as e:
        QMessageBox.critical(None, "Erro", f"Erro ao converter planilha:\n{str(e)}")


def aplicar_formatacao_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    # Estilo de cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste de largura automática
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(caminho)
