import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def exportar_para_excel(df_geral, df_diesel, df_gasolina, df_pneu, caminho):
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        # Aba geral padrão
        df_geral.to_excel(writer, sheet_name='Todos os Itens', index=False)

        # Diesel
        df_diesel.to_excel(writer, sheet_name='Diesel', index=False, startrow=7)
        sheet_diesel = writer.book['Diesel']
        inserir_resumo(sheet_diesel, df_diesel, tipo='Diesel', coef=1.12)

        # Gasolina
        df_gasolina.to_excel(writer, sheet_name='Gasolina', index=False, startrow=7)
        sheet_gasolina = writer.book['Gasolina']
        inserir_resumo(sheet_gasolina, df_gasolina, tipo='Gasolina', coef=1.39)

        df_pneu.to_excel(writer, sheet_name='Pneu', index=False, startrow=7)
        sheet_pneu = writer.book['Pneu']

    ajustar_formatacao_excel(caminho)

def inserir_resumo(sheet, df, tipo, coef):
    quantidade_total = df["Quantidade"].astype(float).sum()
    valor_estimado = quantidade_total * coef

    sheet.merge_cells('A1:F1')
    sheet['A1'] = f"Cálculo do {tipo}"
    sheet['A1'].font = Font(name='Arial', size=14, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')

    resumo = [
        ["Quantidade Total (L)", f"{quantidade_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
        ["Coeficiente de Cálculo", f"{coef:.2f}".replace(".", ",")],
        ["Valor Estimado", f"R$ {valor_estimado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
    ]

    fill = PatternFill("solid", fgColor="f0f0f0")
    bold = Font(name='Arial', bold=True)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    for i, (label, value) in enumerate(resumo, start=3):
        cell_label = sheet[f"A{i}"]
        cell_label.value = label
        cell_label.font = bold
        cell_label.fill = fill
        cell_label.border = border

        cell_val = sheet[f"B{i}"]
        cell_val.value = value
        cell_val.font = Font(name='Arial')
        cell_val.fill = fill
        cell_val.border = border

def ajustar_formatacao_excel(caminho):
    wb = load_workbook(caminho)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for cell in ws[8]:
            if cell.value:
                cell.font = Font(bold=True, name="Arial", size=10)

        for row in ws.iter_rows(min_row=9):
            for cell in row:
                cell.font = Font(name="Arial", size=10)

        for col in ws.columns:
            max_length = 0
            col_index = col[0].column
            col_letter = get_column_letter(col_index)
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            ajustado = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = ajustado

    wb.save(caminho)
